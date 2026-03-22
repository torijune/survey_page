import logging
import hashlib
import json
import io
import csv
from typing import List, Optional, Dict, Any
from uuid import UUID
from datetime import datetime

from ...domain.entities import Response, ResponseItem, Survey, ConditionalLogic, Question, QuestionType
from ...domain.repositories import ResponseRepository, SurveyRepository
from ...api.models import ResponseCreateRequest, ResponseSubmitRequest, ResponseItemRequest

logger = logging.getLogger(__name__)


class ResponseService:
    """응답 서비스"""
    
    def __init__(
        self, 
        response_repository: ResponseRepository,
        survey_repository: SurveyRepository
    ):
        self.response_repository = response_repository
        self.survey_repository = survey_repository
    
    async def start_response(self, request: ResponseCreateRequest) -> Response:
        """응답 시작 (빈 응답 생성)"""
        response = Response(
            survey_id=UUID(request.survey_id),
            ip_address=request.ip_address,
            user_agent=request.user_agent,
        )
        return await self.response_repository.create_response(response)
    
    async def submit_response(
        self, 
        response_id: str, 
        request: ResponseSubmitRequest
    ) -> Response:
        """응답 제출"""
        # 응답 조회
        response = await self.response_repository.get_response_by_id(UUID(response_id), include_items=False)
        if not response:
            raise ValueError("응답을 찾을 수 없습니다.")
        
        # 설문 조회
        survey = await self.survey_repository.get_survey_by_id(response.survey_id, include_details=True)
        if not survey:
            raise ValueError("설문을 찾을 수 없습니다.")
        
        # 설문이 응답을 받을 수 있는지 확인
        if not survey.can_accept_responses():
            raise ValueError("이 설문은 현재 응답을 받지 않습니다.")
        
        # 중복 제출 체크
        if survey.duplicate_prevention and request.user_info:
            user_info_hash = self._hash_user_info(request.user_info)
            is_duplicate = await self.response_repository.check_duplicate_response(
                survey.id, user_info_hash
            )
            if is_duplicate:
                raise ValueError("이미 응답을 제출하셨습니다.")
            
            response.user_info_hash = user_info_hash
            response.user_info_encrypted = self._encrypt_user_info(request.user_info)
        
        # 응답 검증
        self._validate_response(survey, request.items)
        
        # 응답 항목 저장
        items = [
            ResponseItem(
                response_id=response.id,
                question_id=UUID(item.question_id),
                answer_value=item.answer_value,
                answer_text=item.answer_text,
            )
            for item in request.items
        ]
        await self.response_repository.update_response_items(response.id, items)
        
        # 응답 완료 처리
        response.submitted_at = datetime.utcnow()
        response.is_complete = True
        
        return await self.response_repository.update_response(response)
    
    async def update_response_items(
        self, 
        response_id: str, 
        items: List[ResponseItemRequest]
    ) -> Response:
        """응답 항목 업데이트 (중간 저장)"""
        response = await self.response_repository.get_response_by_id(UUID(response_id), include_items=False)
        if not response:
            raise ValueError("응답을 찾을 수 없습니다.")
        
        response_items = [
            ResponseItem(
                response_id=response.id,
                question_id=UUID(item.question_id),
                answer_value=item.answer_value,
                answer_text=item.answer_text,
            )
            for item in items
        ]
        await self.response_repository.update_response_items(response.id, response_items)
        
        return await self.response_repository.get_response_by_id(response.id, include_items=True)
    
    async def get_response(self, response_id: str) -> Optional[Response]:
        """응답 조회"""
        return await self.response_repository.get_response_by_id(UUID(response_id), include_items=True)
    
    async def get_survey_responses(
        self, 
        survey_id: str, 
        include_items: bool = False
    ) -> List[Response]:
        """설문의 응답 목록 조회"""
        return await self.response_repository.get_responses_by_survey_id(
            UUID(survey_id), 
            include_items=include_items,
            only_complete=True
        )
    
    async def get_response_statistics(self, survey_id: str) -> Dict[str, Any]:
        """응답 통계 조회"""
        return await self.response_repository.get_response_statistics(UUID(survey_id))
    
    async def delete_response(self, response_id: str) -> bool:
        """응답 삭제"""
        return await self.response_repository.delete_response(UUID(response_id))
    
    def _get_section_letter(self, section_index: int) -> str:
        """섹션 번호를 A, B, C... 형식으로 변환"""
        return chr(65 + section_index)  # A=65, B=66, C=67...
    
    def _get_question_number(self, survey, question) -> str:
        """문항 번호 반환 (저장된 question_number 우선, 없으면 A1, A2, B1... 형식으로 생성)"""
        # 저장된 question_number가 있으면 우선 사용
        if hasattr(question, 'question_number') and question.question_number:
            return question.question_number
        
        # 섹션 인덱스 찾기
        section_index = -1
        for i, section in enumerate(survey.sections):
            if section.id == question.section_id:
                section_index = i
                break
        
        if section_index == -1:
            return ""
        
        section_letter = self._get_section_letter(section_index)
        
        # 해당 섹션 내에서 문항 인덱스 찾기 (숨겨진 문항 제외)
        section_questions = [q for q in survey.sections[section_index].questions if not getattr(q, 'is_hidden', False)]
        question_index = -1
        for i, q in enumerate(section_questions):
            if q.id == question.id:
                question_index = i
                break
        
        if question_index == -1:
            return ""
        
        return f"{section_letter}{question_index + 1}"
    
    async def generate_csv(self, survey_id: str) -> bytes:
        """CSV 파일 생성"""
        survey = await self.survey_repository.get_survey_by_id(UUID(survey_id), include_details=True)
        if not survey:
            raise ValueError("설문을 찾을 수 없습니다.")
        
        responses = await self.response_repository.get_all_responses_with_items(UUID(survey_id))
        
        # 문항 목록 생성 (숨겨진 문항 제외)
        questions = []
        for section in survey.sections:
            for question in section.questions:
                if not getattr(question, 'is_hidden', False):
                    questions.append(question)
        
        # CSV 생성
        output = io.StringIO()
        writer = csv.writer(output)
        
        # 헤더 및 컬럼 정의 (ID = 일련번호, RID = 응답ID, TIME = 제출시간)
        headers = ["ID", "RID", "TIME"]
        # export_columns: 각 컬럼이 어떤 문항/필드를 나타내는지 메타데이터
        # 타입: ("question", Question) 또는 ("repeatable_input", Question, field_key, row_index, col_index)
        export_columns = []

        # 반복 입력 문항별로 최대 행 수 계산
        repeatable_layout: Dict[str, Dict[str, Any]] = {}
        for q in questions:
            if q.type == QuestionType.REPEATABLE_INPUTS and getattr(q, "repeatable_config", None):
                parts = (q.repeatable_config or {}).get("parts") or []
                field_keys = [
                    part.get("key")
                    for part in parts
                    if isinstance(part, dict) and part.get("type") in ("input", "select")
                ]
                if not field_keys:
                    continue
                max_rows = 1
                q_id_str = str(q.id)
                for resp in responses:
                    for item in resp.items:
                        if str(item.question_id) == q_id_str:
                            value = item.answer_value
                            if isinstance(value, list) and len(value) > max_rows:
                                max_rows = len(value)
                repeatable_layout[q_id_str] = {"field_keys": field_keys, "max_rows": max_rows}

        for q in questions:
            q_id_str = str(q.id)
            layout = repeatable_layout.get(q_id_str)
            if layout:
                field_keys = layout["field_keys"]
                max_rows = layout["max_rows"]
                question_number = self._get_question_number(survey, q)
                if question_number:
                    question_number = question_number.replace("-", "_")
                for row_index in range(1, max_rows + 1):
                    for col_index, field_key in enumerate(field_keys, start=1):
                        suffix = f"{row_index}{col_index}"
                        if question_number:
                            header_title = f"{question_number}_{suffix}"
                        else:
                            header_title = f"{q.title} {suffix}"
                        headers.append(header_title)
                        export_columns.append(("repeatable_input", q, field_key, row_index, col_index))
            else:
                question_number = self._get_question_number(survey, q)
                if question_number:
                    code = question_number.replace("-", "_")
                else:
                    code = f"Q_{q.id}"
                headers.append(code)
                export_columns.append(("question", q))
        writer.writerow(headers)
        
        # 데이터
        for idx, response in enumerate(responses, start=1):
            row = [
                idx,  # ID (일련번호)
                str(response.id),  # RID
                response.submitted_at.isoformat() if response.submitted_at else "",  # TIME
            ]
            
            item_map = {str(item.question_id): item for item in response.items}
            
            for col in export_columns:
                kind = col[0]
                if kind == "question":
                    q = col[1]
                    item = item_map.get(str(q.id))
                    cell = self._format_answer_for_export(q, item)
                    row.append(cell)
                elif kind == "repeatable_input":
                    q, field_key, row_index = col[1], col[2], col[3]
                    item = item_map.get(str(q.id))
                    cell = self._extract_repeatable_field_value(q, item, field_key, row_index)
                    row.append(cell)
            
            writer.writerow(row)
        
        # UTF-8 BOM 추가
        return ("\ufeff" + output.getvalue()).encode("utf-8")
    
    async def generate_xlsx(self, survey_id: str) -> bytes:
        """XLSX 파일 생성"""
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ValueError("openpyxl 패키지가 설치되지 않았습니다.")
        
        survey = await self.survey_repository.get_survey_by_id(UUID(survey_id), include_details=True)
        if not survey:
            raise ValueError("설문을 찾을 수 없습니다.")
        
        responses = await self.response_repository.get_all_responses_with_items(UUID(survey_id))
        statistics = await self.response_repository.get_response_statistics(UUID(survey_id))
        
        # 문항 목록 생성 (숨겨진 문항 제외)
        questions = []
        for section in survey.sections:
            for question in section.questions:
                if not getattr(question, 'is_hidden', False):
                    questions.append(question)
        
        wb = openpyxl.Workbook()
        
        # Sheet 1: Raw Data
        ws1 = wb.active
        ws1.title = "응답 데이터"
        
        # 헤더 및 컬럼 정의 (CSV와 동일한 규칙 사용: ID, RID, TIME)
        headers = ["ID", "RID", "TIME"]
        export_columns = []

        repeatable_layout: Dict[str, Dict[str, Any]] = {}
        for q in questions:
            if q.type == QuestionType.REPEATABLE_INPUTS and getattr(q, "repeatable_config", None):
                parts = (q.repeatable_config or {}).get("parts") or []
                field_keys = [
                    part.get("key")
                    for part in parts
                    if isinstance(part, dict) and part.get("type") in ("input", "select")
                ]
                if not field_keys:
                    continue
                max_rows = 1
                q_id_str = str(q.id)
                for resp in responses:
                    for item in resp.items:
                        if str(item.question_id) == q_id_str:
                            value = item.answer_value
                            if isinstance(value, list) and len(value) > max_rows:
                                max_rows = len(value)
                repeatable_layout[q_id_str] = {"field_keys": field_keys, "max_rows": max_rows}

        for q in questions:
            q_id_str = str(q.id)
            layout = repeatable_layout.get(q_id_str)
            if layout:
                field_keys = layout["field_keys"]
                max_rows = layout["max_rows"]
                question_number = self._get_question_number(survey, q)
                if question_number:
                    question_number = question_number.replace("-", "_")
                for row_index in range(1, max_rows + 1):
                    for col_index, field_key in enumerate(field_keys, start=1):
                        suffix = f"{row_index}{col_index}"
                        if question_number:
                            header_title = f"{question_number}_{suffix}"
                        else:
                            header_title = f"Q_{q.id}_{suffix}"
                        headers.append(header_title)
                        export_columns.append(("repeatable_input", q, field_key, row_index, col_index))
            else:
                question_number = self._get_question_number(survey, q)
                if question_number:
                    code = question_number.replace("-", "_")
                else:
                    code = f"Q_{q.id}"
                headers.append(code)
                export_columns.append(("question", q))
        ws1.append(headers)
        
        # 데이터
        for idx, response in enumerate(responses, start=1):
            row = [
                idx,  # ID (일련번호)
                str(response.id),  # RID
                response.submitted_at.isoformat() if response.submitted_at else "",  # TIME
            ]
            
            item_map = {str(item.question_id): item for item in response.items}
            
            for col in export_columns:
                kind = col[0]
                if kind == "question":
                    q = col[1]
                    item = item_map.get(str(q.id))
                    cell = self._format_answer_for_export(q, item)
                    row.append(cell)
                elif kind == "repeatable_input":
                    q, field_key, row_index = col[1], col[2], col[3]
                    item = item_map.get(str(q.id))
                    cell = self._extract_repeatable_field_value(q, item, field_key, row_index)
                    row.append(cell)
            
            ws1.append(row)
        
        # Sheet 2: Summary (선택지 통계만, 반복 입력/텍스트 문항 제외)
        ws2 = wb.create_sheet("요약 통계")
        ws2.append(["총 응답 수", statistics["total_responses"]])
        ws2.append([])
        
        for q in questions:
            # 텍스트/서술형 문항은 요약 통계에서 제외
            if q.type in [QuestionType.SHORT_TEXT, QuestionType.LONG_TEXT]:
                continue

            # ===== 2-2. 반복 입력 안의 select 파트 통계 =====
            if q.type == QuestionType.REPEATABLE_INPUTS and getattr(q, "repeatable_config", None):
                parts = (q.repeatable_config or {}).get("parts") or []
                select_parts = [p for p in parts if isinstance(p, dict) and p.get("type") == "select"]

                if select_parts:
                    question_number = self._get_question_number(survey, q)
                    question_title = f"{question_number}. {q.title}" if question_number else q.title

                    # 문항 타이틀
                    ws2.append([f"문항: {question_title}"])
                    ws2.append([])

                    # 헤더: 응답 수 / N / %
                    ws2.append(["응답 수", "N", "%"])

                    # 응답 수: 전체 설문 응답 수와 동일하게 사용
                    response_count = statistics["total_responses"]
                    # 각 select 파트마다 통계 계산 (응답자가 해당 유형을 한 번 이상 선택했는지 기준)
                    for s_part in select_parts:
                        field_key = s_part.get("key")
                        part_options = s_part.get("options") or []
                        if not field_key or not part_options:
                            continue

                        # value -> 응답자 수 매핑
                        value_to_respondents: Dict[Any, int] = {}
                        for resp in responses:
                            # 이 응답자가 이 옵션을 한 번이라도 선택했는지 여부를 체크
                            picked_values = set()
                            for item in resp.items:
                                if str(item.question_id) != str(q.id):
                                    continue
                                val = item.answer_value
                                rows = val if isinstance(val, list) else [val] if isinstance(val, dict) else []
                                for row in rows:
                                    if isinstance(row, dict):
                                        raw = row.get(field_key)
                                        if raw is not None:
                                            mapped = self._map_repeatable_select_value(q, field_key, raw)
                                            picked_values.add(mapped)
                            for v in picked_values:
                                value_to_respondents[v] = value_to_respondents.get(v, 0) + 1

                        if response_count > 0:
                            for idx, opt in enumerate(part_options, start=1):
                                mapped_val = self._map_repeatable_select_value(q, field_key, opt.get("value"))
                                count = value_to_respondents.get(mapped_val, 0)
                                percent = round((count / response_count) * 100, 1) if response_count > 0 else 0.0
                                label = f"{idx} {opt.get('label')}"
                                ws2.append([label, count, f"{percent}%"])

                    ws2.append([])
                # select 파트가 없으면 이 문항은 요약 통계에 포함하지 않음
                continue

            # ===== 2-1. 일반 선택 문항 통계 (question_options 기반) =====
            options = getattr(q, "options", None) or []
            if not options:
                continue

            q_stats = statistics["question_stats"].get(str(q.id), {})
            # response_count는 이 문항에 대해 어떤 값이든 있는 응답 수
            response_count = q_stats.get("response_count", 0)
            value_counts = q_stats.get("value_counts", {}) or {}

            if response_count == 0:
                continue

            question_number = self._get_question_number(survey, q)
            question_title = f"{question_number}. {q.title}" if question_number else q.title

            # 문항 타이틀
            ws2.append([f"문항: {question_title}"])
            ws2.append([])

            # 헤더: 응답 수 / N / %
            ws2.append(["응답 수", "N", "%"])

            for idx, opt in enumerate(options, start=1):
                count = value_counts.get(opt.value, 0)
                percent = round((count / response_count) * 100, 1)
                label = f"{idx} {opt.label}"
                ws2.append([label, count, f"{percent}%"])

            ws2.append([])

        # Sheet 3: 변수 설명 (컬럼 코드 ↔ 문항 제목)
        ws3 = wb.create_sheet("변수 설명")
        ws3.append(["변수", "변수 설명"])
        # 공통 변수
        ws3.append(["ID", "일련번호"])
        ws3.append(["RID", "응답 ID"])
        ws3.append(["TIME", "설문 제출시간"])

        # 질문별 변수 코드
        # 먼저 일반 문항 코드
        for q in questions:
            question_number = self._get_question_number(survey, q)
            if question_number:
                code = question_number.replace("-", "_")
            else:
                code = f"Q_{q.id}"
            ws3.append([code, q.title])

        # 반복 입력 세부 변수 코드 (행/열 별)
        for q in questions:
            q_id_str = str(q.id)
            layout = repeatable_layout.get(q_id_str)
            if not layout:
                continue
            field_keys = layout["field_keys"]
            max_rows = layout["max_rows"]
            question_number = self._get_question_number(survey, q)
            if question_number:
                base = question_number.replace("-", "_")
            else:
                base = f"Q_{q.id}"
            for row_index in range(1, max_rows + 1):
                for col_index, _field_key in enumerate(field_keys, start=1):
                    suffix = f"{row_index}{col_index}"
                    code = f"{base}_{suffix}"
                    ws3.append([code, q.title])
        
        # 바이트로 변환
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()

    def _format_answer_for_export(self, question, item) -> str:
        """다운로드용 셀 값 포맷터.

        - answer_text가 있으면 그대로 사용
        - 선택지 value가 'option_1' 같은 형식이면 숫자만 추출해 저장
        - 배열인 경우 각 요소를 같은 규칙으로 변환 후 '1, 2, 3' 형태로 저장
        - dict(리커트, 반복입력 등)는 JSON 문자열로 유지
        """
        if not item:
            return ""

        if item.answer_text:
            return item.answer_text

        value = item.answer_value
        if value is None:
            return ""

        def normalize_option_code(raw):
            # 이미 숫자이거나 숫자 문자열이면 그대로 사용
            if isinstance(raw, (int, float)):
                return raw
            if isinstance(raw, str):
                # option_1, option_2 ... → 숫자 부분만
                if raw.startswith("option_"):
                    suffix = raw.split("_", 1)[1]
                    if suffix.isdigit():
                        return suffix
                    # suffix가 숫자가 아니면 그대로 반환
                    return raw
            return raw

        # 여러 값 (다중 선택 등)
        if isinstance(value, list):
            mapped = [str(normalize_option_code(v)) for v in value]
            return ", ".join(mapped)

        # 객체 값 (리커트, 반복 입력 등)은 그대로 JSON 문자열로
        if isinstance(value, dict):
            return json.dumps(value, ensure_ascii=False)

        # 단일 값
        return str(normalize_option_code(value))

    def _map_repeatable_select_value(self, question: Question, field_key: str, raw: Any) -> Any:
        """반복 입력 select 파트일 경우, 옵션 순서 기반 1,2,... 인덱스로 변환."""
        if raw is None:
            return ""
        if not getattr(question, "repeatable_config", None):
            return raw
        parts = (question.repeatable_config or {}).get("parts") or []
        for part in parts:
            if isinstance(part, dict) and part.get("type") == "select" and part.get("key") == field_key:
                options = part.get("options") or []
                for idx, opt in enumerate(options, start=1):
                    if opt.get("value") == raw:
                        return idx
                return raw
        return raw

    def _extract_repeatable_field_value(self, question: Question, item, field_key: str, row_index: int) -> str:
        """반복 입력 문항에서 특정 행과 입력칸(field_key)의 값을 추출.

        - answer_value가 리스트이면 row_index(1-based)에 해당하는 행에서 key 값을 사용
        - answer_value가 dict이면 단일 행으로 보고 그 dict에서 바로 key 값을 사용
        """
        if not item or not field_key:
            return ""

        value = item.answer_value
        if value is None:
            return ""

        if isinstance(value, list) and value:
            idx = row_index - 1
            if 0 <= idx < len(value):
                    row_val = value[idx]
                    if isinstance(row_val, dict):
                        v = row_val.get(field_key)
                        mapped = self._map_repeatable_select_value(question, field_key, v)
                        return "" if mapped is None else str(mapped)
            # 지정한 행이 없으면 빈 문자열
            return ""

        if isinstance(value, dict):
            v = value.get(field_key)
            mapped = self._map_repeatable_select_value(question, field_key, v)
            return "" if mapped is None else str(mapped)

        # 그 외 형태는 문자열로 반환
        return str(value)
    
    def _is_question_visible_for_response(
        self,
        question: Question,
        item_map: Dict[str, ResponseItemRequest],
    ) -> bool:
        """조건문이 있으면 제출된 응답 기준으로 해당 문항이 '보였어야 하는지' 판단. 조건 없으면 True."""
        conditions: Optional[List[ConditionalLogic]] = question.conditional_logic
        if not conditions:
            return True

        def evaluate_one(c: ConditionalLogic) -> bool:
            item = item_map.get(str(c.question_id))
            if not item:
                return False
            answer_value = item.answer_value
            condition_values = c.value if isinstance(c.value, list) else [c.value]
            if condition_values is None:
                condition_values = []

            if c.operator == "equals":
                if isinstance(answer_value, list):
                    return any(v in condition_values for v in answer_value)
                return answer_value in condition_values
            if c.operator == "not_equals":
                if isinstance(answer_value, list):
                    return not any(v in condition_values for v in answer_value)
                return answer_value not in condition_values
            if c.operator == "contains":
                if isinstance(answer_value, list):
                    return any(v in answer_value for v in condition_values)
                return answer_value in condition_values
            if c.operator == "not_contains":
                if isinstance(answer_value, list):
                    return not any(v in answer_value for v in condition_values)
                return answer_value not in condition_values
            if c.operator == "greater_than":
                try:
                    num_ans = float(answer_value) if answer_value is not None else float("nan")
                except (TypeError, ValueError):
                    return False
                if num_ans != num_ans:
                    return False
                for v in condition_values:
                    if v is None:
                        continue
                    try:
                        if num_ans > float(v):
                            return True
                    except (TypeError, ValueError):
                        pass
                return False
            if c.operator == "less_than":
                try:
                    num_ans = float(answer_value) if answer_value is not None else float("nan")
                except (TypeError, ValueError):
                    return False
                if num_ans != num_ans:
                    return False
                for v in condition_values:
                    if v is None:
                        continue
                    try:
                        if num_ans < float(v):
                            return True
                    except (TypeError, ValueError):
                        pass
                return False
            return False

        all_met = all(evaluate_one(c) for c in conditions)
        action = conditions[0].action if conditions else "show"
        return all_met if (action == "show") else (not all_met)

    def _validate_response(self, survey: Survey, items: List[ResponseItemRequest]) -> None:
        """응답 검증. 조건문이 있는 문항은 조건을 만족한 응답자에게만 필수로 적용."""
        item_map = {item.question_id: item for item in items}

        for section in survey.sections:
            for question in section.questions:
                item = item_map.get(str(question.id))

                # 필수 항목: 조건문이 있으면 '보였어야 하는 경우'에만 필수, 없으면 항상 필수
                is_visible = self._is_question_visible_for_response(question, item_map)
                if question.required and is_visible:
                    if not item or (item.answer_value is None and not item.answer_text):
                        raise ValueError(f"필수 항목입니다: {question.title}")

                # 검증 규칙 체크
                if item and question.validation_rules:
                    rules = question.validation_rules
                    
                    if item.answer_text:
                        text_len = len(item.answer_text)
                        if rules.min_length and text_len < rules.min_length:
                            raise ValueError(f"최소 {rules.min_length}자 이상 입력해주세요: {question.title}")
                        if rules.max_length and text_len > rules.max_length:
                            raise ValueError(f"최대 {rules.max_length}자까지 입력 가능합니다: {question.title}")
                        
                        if rules.pattern == "email":
                            import re
                            if not re.match(r"[^@]+@[^@]+\.[^@]+", item.answer_text):
                                raise ValueError(f"올바른 이메일 형식이 아닙니다: {question.title}")
                    
                    if item.answer_value is not None and isinstance(item.answer_value, (int, float)):
                        if rules.min_value is not None and item.answer_value < rules.min_value:
                            raise ValueError(f"최소값은 {rules.min_value}입니다: {question.title}")
                        if rules.max_value is not None and item.answer_value > rules.max_value:
                            raise ValueError(f"최대값은 {rules.max_value}입니다: {question.title}")
    
    def _hash_user_info(self, user_info: str) -> str:
        """사용자 정보 해시 생성"""
        return hashlib.sha256(user_info.encode()).hexdigest()
    
    def _encrypt_user_info(self, user_info: str) -> str:
        """사용자 정보 암호화 (간단한 구현)"""
        # TODO: 실제 암호화 구현 (cryptography 라이브러리 사용)
        import base64
        return base64.b64encode(user_info.encode()).decode()

